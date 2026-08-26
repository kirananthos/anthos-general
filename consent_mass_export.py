import pandas as pd, requests, re
import subprocess
import openpyxl
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from concurrent.futures import ThreadPoolExecutor
import sys

def downloadFile(entry):
    fname, url = entry
    url = url.strip()
    try:
        html = requests.get(url).text
    except requests.exceptions.RequestException as e:
        with open("errors.txt", "a") as f:
            f.write(f"{fname}: FAILED to fetch {url} for file name: {fname}({e})\n")
        return
    soup = BeautifulSoup(html, "html.parser")

    # download friendly record if necessary
    link = soup.find("a", string=lambda s: s and s.strip() == "Print Friendly Record")
    if link:
        html = requests.get(urljoin(url, link["href"])).text

    subprocess.run(["wkhtmltopdf", "-", f"{fname}.pdf"], input=html.encode())


# takes excel as argument to script
def main():
    excel_path = sys.argv[1]

    # use openpyxl directly so we get real hyperlink targets, not just the
    # display text pandas would give us for cells that are clickable links
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    # builds a list of [[filename, url], [filename2, url2]...]
    # file name = ContactID_ParticipantName_ConsentType_Consent.html
    names_and_urls = []
    for row in ws.iter_rows(min_row=2):
        r = dict(zip(headers, row))
        for col in headers:
            if col in ("Contact ID", "Full Name"):
                continue
            cell = r[col]
            url = cell.hyperlink.target if cell.hyperlink else cell.value
            if not url:
                continue
            fname = re.sub(r"[^A-Za-z0-9._-]", "_", f"{r['Contact ID'].value}_{r['Full Name'].value}_{col}_Consent")
            names_and_urls.append((fname, url))

    # batch http requests
    pool = ThreadPoolExecutor(max_workers=20)
    list(pool.map(downloadFile, names_and_urls))
    pool.shutdown()


if __name__ == "__main__":
    main()